import datetime
import os
import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import urllib.request
import json
from openpyxl import Workbook, load_workbook

request_path = r'https://stock.finance.sina.com.cn/forex/api/openapi.php/ForexService.getBankForexList?from=SGD&_=1724813755320'

class Email:
    def __init__(self):
        self.smtp = smtplib.SMTP()
        self.smtp.connect('')
        self.email_body = MIMEMultipart('mixed')
        self.email_address = ''
        self.email_auth_code = ''
        self.smtp.login(self.email_address, self.email_auth_code)

    def send_email(self, receiver_email, email_title, email_content, attachment_path, files):
        self.email_body['Subject'] = email_title
        self.email_body['From'] = self.email_address
        self.email_body['To'] = receiver_email

        for file in files:
            file_path = os.path.join(attachment_path, file)
            if os.path.isfile(file_path):
                att = MIMEText(open(file_path, 'rb').read(), 'base64', 'utf-8')
                att["Content-Type"] = 'application/octet-stream'
                att.add_header("Content-Disposition", "attachment", filename=("gbk", "", file))
                self.email_body.attach(att)

        text_plain = MIMEText(email_content, 'plain', 'utf-8')
        self.email_body.attach(text_plain)
        self.smtp.sendmail(self.email_address, receiver_email, self.email_body.as_string())

class Bank:
    def __init__(self, bank_name, bank_buy_price, bank_sell_price):
        self.bank_name = bank_name
        self.bank_buy_price = bank_buy_price
        self.bank_sell_price = bank_sell_price
        self.time = 0

class request_Trans:
    def __init__(self):
        self.request_path = request_path
        self.time = 0

    def requestInfo(self):
        request = urllib.request.Request(url=self.request_path)
        response = urllib.request.urlopen(request)
        content = response.read().decode('utf-8')
        content_json = json.loads(content)
        result = content_json['result']
        data = result['data']
        result = []
        for exact_bank_data in data:
            if self.time == 0:
                self.time = exact_bank_data['updatetime']
            bank_account = Bank(exact_bank_data['bank'], exact_bank_data['xh_buy_price'], exact_bank_data['xh_sell_price'])
            result.append(bank_account)
        return result

class Excel_Processor:
    def __init__(self, file_path):
        self.contact_info = dict()
        self.original_file_path = file_path
        if os.path.isfile(file_path):
            self.df = load_workbook(file_path)
        else:
            self.df = Workbook()
            self.df.active.title = "汇率数据"
            self.df.active.append(["银行", "买入价格", "卖出价格", "更新时间"])  # 添加表头

        self.ws = self.df.active

    def write_excel(self, row, col, content):
        self.ws.cell(row=row, column=col).value = content

    def save_excel(self):
        self.df.save(self.original_file_path)

print("开始请求数据")
new_re = request_Trans()
list_bank_account = new_re.requestInfo()
print("请求成功")
start_row = 2

excel_file = Excel_Processor('getInfo_money.xlsx')
for bank_account in list_bank_account:
    excel_file.write_excel(start_row, 1, bank_account.bank_name)
    excel_file.write_excel(start_row, 2, bank_account.bank_buy_price)
    excel_file.write_excel(start_row, 3, bank_account.bank_sell_price)
    start_row += 1

excel_file.write_excel(3, 4, new_re.time)

print("编辑成功")
excel_file.save_excel()
email = Email()
email.send_email("", new_re.time, "更新的汇率，请查收", r'', ['getInfo_money.xlsx'])
input("成功发送，请按回车键退出")